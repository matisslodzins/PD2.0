import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import ElementNotInteractableException
from selenium.common.exceptions import NoSuchElementException
import time
from openpyxl import Workbook, load_workbook 

#parametru ievade
minCena = input("ievadiet minimālo cenu: ")
maxCena = input("ievadiet maksimālo cenu: ")
minGads = input("ievadiet minimālo vecumu: ")
maxGads = input("ievadiet maksimalo vecumu: ")

print(" ")

print("1. Benzīns, 2. Benzīns/gāze, 3. Dīzelis, 4. Hybrid, 5. Elekstriskais")
dzinejs = input("ievadiet dzinēja tipu skaitli: ")
if dzinejs =="1":
    dzinejs =="Benzīns"
elif dzinejs == "2":
    dzinejs = "Benzīns/gāze"
elif dzinejs == "3":
    dzinejs = " Dīzelis"  
elif dzinejs == "4":
    dzinejs = "Hybrid" 
elif dzinejs == "5":
    dzinejs = " Elekstriskais" 
else: print("kļūda, neeksistē")
dzinejs = dzinejs.lstrip().rstrip()

print(" ")

karba = input("ievadiet 1, ja vēlaties automātu un 2, ja vēlaties manuālu: ")
if karba == 1:
    karba = "Automāts"
else: karba = "Manuāla"
karba = karba.lstrip().rstrip()

print(' ')

print("1. Apvidus, 2. Hečbeks, 3. Kabriolets, 4. Mikroautobuss,")
print("5. Minivens, 6. Pikaps, 7. Sedans, 8. Kupeja, 9. Universāls ")
tips = input("ievadiet auto tipa skaitli: ")
if tips == "1":
    tips =="Apvidus"
elif tips == "2":
    tips = "Hečbeks"
elif tips == "3":
    tips = "Kabriolets"  
elif tips == "4":
    tips = "Mikroautobuss" 
elif tips == "5":
    tips = "Minivens" 
elif tips == "6":
    tips = "Pikaps" 
elif tips == "7":
    tips = "Sedans" 
elif tips == "8":
    tips = "Kupeja"    
elif tips == "9":
    tips = "Universāls" 
else: print("kļūda, neeksistē")  
tips = tips.lstrip().rstrip()

#minCena = 2000
#maxCena = 4000
#minGads = 2000
#maxGads = 2002
#dzinejs = "Dīzelis"
#karba = "Manuāla"
#tips = "Kupeja"

#servera un faila atvēršana
service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)
url = "https://www.ss.lv/lv/transport/cars/"
driver.get(url)
time.sleep(2)

wb=load_workbook('masinu_gramata.xlsx')
ws=wb.active

#coocies
find = driver.find_element(By.XPATH, '//*[@id="cookie_confirm_body"]/table/tbody/tr/td[2]/button')
find.click()
#minimālās cenas atzīmēšana
find = driver.find_element(By.ID, "f_o_8_min")
find.send_keys(minCena)
#minimālās cenas atzīmēšana
find = driver.find_element(By.ID, "f_o_8_max")
find.send_keys(maxCena)
time.sleep(1)
#minimālā gada atzīmēšana
find = driver.find_element(By.XPATH, "//select[@name='topt[18][min]']")
find.click()
find = driver.find_element(By.XPATH, f"//select[@name='topt[18][min]']/option[text()='{minGads}']")
find.click()
#maksimālā gada atzīmēšana
find = driver.find_element(By.XPATH, "//select[@name='topt[18][max]']")
find.click()
find = driver.find_element(By.XPATH, f"//select[@name='topt[18][max]']/option[text()='{maxGads}']")
find.click()
#dzinēja izvēle
find = driver.find_element(By.XPATH, "//select[@name='opt[34]']")
find.click()
find = driver.find_element(By.XPATH, f"//select[@name='opt[34]']/option[text()='{dzinejs}']")
find.click()
#ātruma kārbas izvēle
find = driver.find_element(By.XPATH, "//select[@name='opt[35]']")
find.click()
find = driver.find_element(By.XPATH, f"//select[@name='opt[35]']/option[text()='{karba}']")
find.click()
#tipa izvēle
find = driver.find_element(By.XPATH, "//select[@name='opt[32]']")
find.click()
find = driver.find_element(By.XPATH, f"//select[@name='opt[32]']/option[text()='{tips}']")
find.click()

#clear iepriekšējo
for row in ws['A1:H5000']:
    for cell in row:
        cell.value = None

#headlines
ws['a'+str(1)].value = 'Mašīna'
ws['b'+str(1)].value = 'Izlaidums'
ws['c'+str(1)].value = 'Dzinējs'
ws['d'+str(1)].value = 'Kārba'
ws['e'+str(1)].value = 'Tips'
ws['f'+str(1)].value = 'Nobraukums'
ws['g'+str(1)].value = 'Cena'
ws['h'+str(1)].value = 'Links'

masinas = driver.find_elements(By.XPATH, '//table[@id="page_main"]//tr[@id="head_line"]/following-sibling::tr')

col = 2

for i in range(len(masinas)):
    try:     
        # programma apstajas, ja neredz kur spiest, sis scrollo
        driver.execute_script("arguments[0].scrollIntoView();", masinas[i])
        masinas[i].click()
        #seit lasis visu info un liks tabula
        
        find_marka = driver.find_element(By.ID, "tdo_31")    #marka
        fmarka = find_marka.text
        ws['a'+str(col)].value = fmarka

        find_gads = driver.find_element(By.ID, "tdo_18")    #gads
        fgads = find_gads.text
        ws['B'+str(col)].value = fgads

        find_dzinejs = driver.find_element(By.ID, "tdo_15")    #dzinejs
        fdzinejs = find_dzinejs.text
        ws['C'+str(col)].value = fdzinejs

        find_karba = driver.find_element(By.ID, "tdo_35")    #kārba
        fkarba = find_karba.text
        ws['D'+str(col)].value = fkarba

        #dazam masinam nav nobraukums, tādēļ nepieciešams cikls
        try:
            find_nobraukums = driver.find_element(By.ID, "tdo_16")   #nobraukums
            fnobraukums = find_nobraukums.text

        except NoSuchElementException:
            fnobraukums = "Nav norādīts"
        
        ws['f'+str(col)].value = fnobraukums

        find_uzbuve = driver.find_element(By.ID, "tdo_32")    #uzbuves tips
        fuzbuve = find_uzbuve.text
        ws['E'+str(col)].value = fuzbuve

        find_cena = driver.find_element(By.ID, "tdo_8")    #cena
        fcena = find_cena.text
        ws['g'+str(col)].value = fcena

        find_links = driver.current_url  #links
        ws['h'+str(col)].value = find_links

        col = col+1        
        time.sleep(1)
        
        #atpakal uz lapu
        driver.back()
        time.sleep(1)
        table = driver.find_element(By.XPATH, '//table')
        masinas = table.find_elements(By.XPATH, './/tr[@id="head_line"]/following-sibling::tr')

    except ElementNotInteractableException:
        break
    
wb.save('masinu_gramata.xlsx')
wb.close()
